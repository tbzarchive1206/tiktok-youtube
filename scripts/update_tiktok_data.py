"""Build data.js from the public Google Sheet every six hours."""
from __future__ import annotations
import io,json,re,urllib.request
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

SHEET_ID="1C0DP7DKN5QCO5GXdNDWYmvuK8RtGEkYp"
URL=f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/"data.js"
ALLOWED={"istent_theboyz","theboyz_officl","jakeybaee2","kebean.moon","eric.sohn22"}
ALIASES={
 "date":{"data","date","upload date","data dodania"},
 "description":{"opis tiktoka","opis","description","caption"},
 "hashtags":{"hashtagi","hashtags","tags"},
 "members":{"czlonkowie","członkowie","members","member"},
 "tiktok":{"link tiktok","link do tiktoka","tiktok url","tiktok link","tiktok"},
 "youtube":{"link youtube","link do youtube","youtube","youtube url","youtube link","link youtube shorts"},
 "drive":{"link google drive","google drive","google drive link","drive","drive link","link drive","link do google drive","link do pobrania","download link"}
}

def norm(v):return re.sub(r"\s+"," ",str(v or "").strip().lower().replace("_"," "))
def cell_value(c):return (str(c.hyperlink.target).strip() if c.hyperlink and c.hyperlink.target else str(c.value or "").strip())
def cols(headers):
 h=[norm(x) for x in headers];o={}
 for key,aliases in ALIASES.items():
  for i,v in enumerate(h):
   if v in aliases:o[key]=i;break
 return o

def tid(url):
 m=re.search(r"/video/(\d+)",url);return m.group(1) if m else ""
def account(url):
 m=re.search(r"tiktok\.com/@([^/?#]+)",url,re.I);return m.group(1) if m else ""
def parsedate(v,id_):
 if isinstance(v,datetime):d=v
 else:
  s=re.sub(r"\D","",str(v or ""));d=None
  for fmt,l in (("%y%m%d",6),("%Y%m%d",8)):
   if len(s)==l:
    try:d=datetime.strptime(s,fmt);break
    except ValueError:pass
  if d is None and id_.isdigit():
   try:d=datetime.utcfromtimestamp(int(id_)>>32)
   except Exception:pass
 return (d.strftime("%Y-%m-%d"),d.year,d.strftime("%y%m%d")) if d else ("",0,"")
def memberlist(v):return [x.strip() for x in re.split(r"[,;/|]+",str(v or "")) if x.strip()]

def main():
 req=urllib.request.Request(URL,headers={"User-Agent":"tbzarchive-github-sync/3.0"})
 with urllib.request.urlopen(req,timeout=60) as r:data=r.read()
 if not data.startswith(b"PK"):raise RuntimeError("Google Sheet is not publicly downloadable as XLSX.")
 wb=load_workbook(io.BytesIO(data),data_only=True,read_only=False)
 items=[]
 for ws in wb.worksheets:
  rows=list(ws.iter_rows())
  header_i=None;cmap=None
  for i,row in enumerate(rows[:20]):
   c=cols([x.value for x in row])
   if "tiktok" in c:
    header_i=i;cmap=c;break
  if header_i is None:continue
  for row in rows[header_i+1:]:
   def get(k):
    i=cmap.get(k);return cell_value(row[i]) if i is not None and i<len(row) else ""
   tt=get("tiktok");id_=tid(tt);acc=account(tt)
   if not id_ or acc not in ALLOWED:continue
   date,year,code=parsedate(get("date"),id_)
   items.append({"tiktokId":id_,"account":acc,"date":date,"dateCode":code,"year":year,"description":get("description"),"hashtags":get("hashtags"),"members":memberlist(get("members")),"tiktokUrl":tt,"youtubeUrl":get("youtube"),"driveUrl":get("drive")})
 unique={x["tiktokId"]:x for x in items};items=list(unique.values());items.sort(key=lambda x:(x["date"],x["tiktokId"]),reverse=True)
 if not items:raise RuntimeError("No TikTok rows found. data.js was not changed.")
 payload={"sourceSpreadsheet":f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit","videos":items}
 text=json.dumps(payload,ensure_ascii=False,separators=(",",":"),).replace("<","\\u003c").replace(">","\\u003e").replace("&","\\u0026")
 OUT.write_text("window.TIKTOK_ARCHIVE_DATA="+text+";\n",encoding="utf-8")
 print(f"Wrote {len(items)} videos; YouTube links: {sum(bool(x['youtubeUrl']) for x in items)}")
if __name__=="__main__":main()
